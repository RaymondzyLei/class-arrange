# -*- coding: utf-8 -*-
"""把开课 Excel 转换为前端可用的 TS 数据文件（可复用）。

用法:
    uv run python scripts/excel_to_ts.py [xlsx路径]

不传路径时自动取项目根目录下唯一的 *.xlsx。
输出:
    src/data/courses.ts    数据（导出 const courses）
    src/data/courses.d.ts  类型定义
"""
from __future__ import annotations

import glob
import json
import os
import re
import sys
from collections import Counter

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_TS = os.path.join(ROOT, "src", "data", "courses.ts")


# ---------- 字段解析 ----------
DEPT_RE = re.compile(r"^\s*(\S+)\s+(.+?)\s*$")
# 时间地点片段: "1(8,9)" 或 "1(8,9,10)" 等
SLOT_RE = re.compile(r"(\d+)\s*\((\d+(?:,\d+)*)\)")
# 周次头部: 匹配 "1~9周" / "1周" / "1,3,5周" / "2,7~9(单)周" / "8~10(双)周" 直到第一个 "周"
WEEK_HEAD_RE = re.compile(r"^([0-9~,~()单双]+周)")


def to_int(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return default


def to_num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def split_dept(s):
    if not s:
        return {"code": "", "name": ""}
    m = DEPT_RE.match(str(s))
    if m:
        return {"code": m.group(1), "name": m.group(2)}
    return {"code": "", "name": str(s).strip()}


def split_classes(s):
    if not s:
        return []
    parts = re.split(r"[,，]", str(s))
    return [p.strip() for p in parts if p.strip()]


def parse_weeks(week_str):
    """解析周次头部（不含"周"字），返回 [start,end] 闭区间列表。

    支持: "1~9" / "1" / "1,3,5" / "2,7~9(单)" / "8~10(双)" / "1,9~14"
    单/双周: 在 range 上过滤奇/偶周。
    """
    ranges = []
    for part in week_str.split(","):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"(\d+)\s*~\s*(\d+)(?:\((单|双)\))?", part)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            parity = m.group(3)
            if parity == "单":
                ranges.append([w for w in range(a, b + 1) if w % 2 == 1])
            elif parity == "双":
                ranges.append([w for w in range(a, b + 1) if w % 2 == 0])
            else:
                ranges.append([a, b])
        else:
            m2 = re.match(r"(\d+)", part)
            if m2:
                ranges.append([int(m2.group(1)), int(m2.group(1))])
    return ranges


def parse_schedule(raw):
    """解析 '时间地点' 字段为 ScheduleSlot[]。

    每段格式: "<周次> 教室: 星期(节次) [星期(节次)...]"
    周次示例: 1~9周 / 1周 / 1,3,5周 / 2,7~9(单)周
    多段之间用换行或分号分隔；一段内可有多个教室用分号分隔。
    _x000d_ 是 Excel 中 \\r 的残留，先清理。
    返回 (slots, ok)。
    """
    if not raw:
        return [], True
    s = str(raw).replace("_x000d_", "").strip()
    slots = []
    ok = True
    # 按换行切成"周次段"
    for line in re.split(r"[\n]+", s):
        line = line.strip()
        if not line:
            continue
        hm = WEEK_HEAD_RE.match(line)
        if not hm:
            ok = False
            continue
        week_part = hm.group(1)[:-1]  # 去掉"周"
        week_ranges = parse_weeks(week_part)
        if not week_ranges:
            ok = False
            continue
        rest = line[hm.end():].strip()
        # 一段内多个教室用分号分隔: "5407: 4(11,12); 5507: 1(3,4)"
        for room_seg in rest.split(";"):
            room_seg = room_seg.strip()
            if not room_seg:
                continue
            room = ""
            time_str = room_seg
            colon = room_seg.find(":")
            if colon != -1:
                room = room_seg[:colon].strip()
                time_str = room_seg[colon + 1:].strip()
            else:
                # 无冒号: 取首个非时间 token 当教室
                tokens = room_seg.split()
                time_tokens = []
                for t in tokens:
                    if SLOT_RE.fullmatch(t):
                        time_tokens.append(t)
                    elif not room:
                        room = t
                    else:
                        time_tokens.append(t)
                time_str = " ".join(time_tokens)
            for sm in SLOT_RE.finditer(time_str):
                day = int(sm.group(1))
                periods = [int(x) for x in sm.group(2).split(",")]
                for wr in week_ranges:
                    slots.append({
                        "weeks": wr if len(wr) > 2 else [min(wr), max(wr)],
                        "room": room,
                        "day": day,
                        "periods": periods,
                    })
            if not SLOT_RE.search(time_str):
                ok = False
    return slots, ok


# ---------- 主流程 ----------
def main():
    if len(sys.argv) > 1:
        xlsx = sys.argv[1]
    else:
        xlsxs = glob.glob(os.path.join(ROOT, "*.xlsx"))
        if len(xlsxs) != 1:
            print(f"错误: 项目根下找到 {len(xlsxs)} 个 xlsx，请显式指定路径", file=sys.stderr)
            sys.exit(1)
        xlsx = xlsxs[0]
    print(f"读取: {xlsx}")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    data = rows[1:]
    idx = {h: i for i, h in enumerate(headers)}

    def get(r, name):
        i = idx.get(name)
        if i is None:
            return None
        return r[i] if i < len(r) else None

    courses = []
    fail_count = 0
    fail_samples = []
    seen_ids = Counter()
    for r in data:
        cid = str(get(r, "课堂号") or "").strip()
        seen_ids[cid] += 1
        raw_sched = get(r, "时间地点")
        slots, ok = parse_schedule(raw_sched)
        if not ok:
            fail_count += 1
            if len(fail_samples) < 20:
                fail_samples.append((cid, str(get(r, "课程名")), repr(raw_sched)))
        courses.append({
            "id": cid,
            "courseName": str(get(r, "课程名") or "").strip(),
            "department": split_dept(get(r, "开课单位")),
            "teacher": str(get(r, "授课教师") or "").strip(),
            "credits": to_num(get(r, "学分")),
            "hours": to_int(get(r, "学时")),
            "level": str(get(r, "学历层次") or "").strip(),
            "sectionType": str(get(r, "课堂类型") or "").strip(),
            "category": str(get(r, "课程范畴分类") or "").strip(),
            "courseType": str(get(r, "课程类型") or "").strip(),
            "language": str(get(r, "授课语言") or "").strip(),
            "examType": str(get(r, "考核方式") or "").strip(),
            "undergradShared": to_int(get(r, "本研同堂")) == 1,
            "enrolled": to_int(get(r, "选课人数")),
            "capacity": to_int(get(r, "限选人数")),
            "classes": split_classes(get(r, "上课班级")),
            "rawSchedule": str(raw_sched or "").strip(),
            "schedule": slots,
        })

    # 写 TS
    os.makedirs(os.path.dirname(OUT_TS), exist_ok=True)
    with open(OUT_TS, "w", encoding="utf-8") as f:
        f.write("// AUTO-GENERATED by scripts/excel_to_ts.py — do not edit by hand.\n")
        f.write("import type { CourseSection } from '@/types';\n\n")
        f.write("export const courses: CourseSection[] = ")
        f.write(json.dumps(courses, ensure_ascii=False, indent=2))
        f.write(";\n")

    # 重复 id 警告
    dup_ids = {k: v for k, v in seen_ids.items() if v > 1 and k}
    print(f"生成 {len(courses)} 条课程 -> {OUT_TS}")
    print(f"类型定义: src/types/index.ts (手写维护，脚本不再生成)")
    print(f"时间地点解析失败行数: {fail_count}")
    if fail_samples:
        print("失败样本（前20）:")
        for cid, name, raw in fail_samples:
            print(f"  {cid} {name}: {raw}")
    if dup_ids:
        print(f"警告: 重复课堂号 {len(dup_ids)} 个:", list(dup_ids.items())[:10])
    else:
        print("课堂号唯一性: 全部唯一")


if __name__ == "__main__":
    main()
